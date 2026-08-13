#!/usr/bin/env python3
"""Ingestion adapter: spreadsheet -> normalised JSON.

This is the only file that knows anything about the spreadsheet. Swapping to a
direct bank connection later means writing a sibling adapter that emits the same
JSON shape, and changing nothing else in the app.

Usage:
    python3 scripts/export-from-workbook.py <workbook.xlsx> [out.json]
"""
import json
import sys
import datetime as dt
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

COL = dict(date=2, desc=3, group=4, category=5, amount=6,
           account=7, mask=8, institution=9, txid=12)


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = Path(sys.argv[1])
    out = Path(sys.argv[2] if len(sys.argv) > 2 else "data/transactions.json")
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(src, data_only=True)
    if "Transactions" not in wb.sheetnames:
        sys.exit(f"no Transactions sheet in {src}")
    ws = wb["Transactions"]

    rows, skipped = [], 0
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, COL["date"]).value
        amt = ws.cell(r, COL["amount"]).value
        if not isinstance(d, dt.datetime) or not isinstance(amt, (int, float)):
            skipped += 1
            continue
        rows.append({
            "externalId": str(ws.cell(r, COL["txid"]).value or f"row-{r}"),
            "date": d.strftime("%Y-%m-%d"),
            "amount": round(float(amt), 2),
            "description": str(ws.cell(r, COL["desc"]).value or "").strip(),
            "category": (str(ws.cell(r, COL["category"]).value).strip()
                         if ws.cell(r, COL["category"]).value else None),
            "group": (str(ws.cell(r, COL["group"]).value).strip()
                      if ws.cell(r, COL["group"]).value else None),
            "account": str(ws.cell(r, COL["account"]).value or "").strip(),
            "mask": str(ws.cell(r, COL["mask"]).value or "").strip(),
            "institution": str(ws.cell(r, COL["institution"]).value or "").strip(),
        })

    cats = []
    if "Categories" in wb.sheetnames:
        cs = wb["Categories"]
        for r in range(2, cs.max_row + 1):
            name = cs.cell(r, 1).value
            if not name:
                continue
            cats.append({
                "name": str(name).strip(),
                "group": str(cs.cell(r, 2).value or "").strip(),
                "type": str(cs.cell(r, 3).value or "").strip(),
            })

    payload = {
        "source": src.name,
        "exportedAt": dt.datetime.now().isoformat(timespec="seconds"),
        "categories": cats,
        "transactions": rows,
    }
    out.write_text(json.dumps(payload, indent=1))
    dates = [r["date"] for r in rows]
    print(f"wrote {out}")
    print(f"  transactions : {len(rows)}  ({min(dates)} .. {max(dates)})")
    print(f"  categories   : {len(cats)}")
    print(f"  skipped rows : {skipped}")


if __name__ == "__main__":
    main()
